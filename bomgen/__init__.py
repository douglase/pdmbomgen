#!/usr/bin/env python3
# SPDX-License-Identifier: MIT
# SPDX-FileCopyrightText: 2026 E Douglas, University of Arizona and contributors
"""bomgen — SolidWorks PDM Professional CSV BOM -> human-readable Excel + HTML.

See BOMGEN_DESIGN.md. Single-module by design; deps: openpyxl (+ stdlib
tomllib). Installable (pyproject.toml at repo root) so downstream "vault"
repos can `pip install` it straight from this repo instead of vendoring
the file — see template-repo/ for the pattern (design decision D7).
"""
from __future__ import annotations

import argparse
import csv
import html
import json
import re
import sys
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from urllib.parse import quote

try:
    import tomllib
except ModuleNotFoundError:  # pragma: no cover (py<3.11)
    tomllib = None

__version__ = "0.1.0"

# --------------------------------------------------------------------------- config

DEFAULT_CONFIG = {
    "project": {
        "title_line": "BOM DRAFT",
        "project_box": "PROJECT NAME AND TITLE BOX",
        "system_name": "",
        "assembly_number": "TOP-LEVEL-00",
        "assembly_name": "Top Level Assembly",
        "assembly_description": "",
        "contact_name": "",
        "contact_info": "",
    },
    "columns": {
        "level": "Level",
        "qty": "Qty",
        "number": "Number",
        "name": "Name",
        "config": "Config",
        "rev": "Rev",
        "description": "Description",
        "found_in": "Found In",
        "cots": "COTS",
        "material": "Material",
        "specs": "Specs",
        "passthrough": [],
    },
    "rules": {
        "part_number_pattern": r"^([A-Za-z]+-[A-Za-z]{2}\d+)",
        "dash_source": "rev",  # reserved for O1
        "min_marker_columns": 5,
        "state_from_found_in": True,
    },
    "output": {"excel_font": "Aptos Narrow"},
    "links": {
        # URL template for linking each filename to a PDM web viewer.
        # Placeholders (either/both), URL-encoded on substitution:
        #   "{found_in}" -> the Found In vault path, local prefix stripped
        #                   (see found_in_strip) and backslashes -> "/".
        #   "{file}"     -> the SolidWorks filename, e.g. NCC-FA001.SLDASM.
        # Empty -> filenames are not linked.
        "file_url_template": "",
        # Local prefix removed from Found In before substituting {found_in},
        # e.g. "D:\\Steward_Obs_PDM" so the web URL base can supply the vault
        # root. A leading drive letter is dropped automatically regardless.
        "found_in_strip": "",
        # URL template for linking a spec document on the budget dashboard;
        # "{spec}" -> the (URL-encoded) spec reference. Empty -> no links.
        "spec_url_template": "",
    },
    "materials": {
        # Enrich rows with properties from a committed materials-database
        # export (the raw /export/raw-json dump: a JSON array of material
        # documents). Local file only — bomgen never touches the network.
        "enabled": False,
        "cache_file": "",              # path to the raw-json dump
        "properties": [],              # DB property keys to show as columns
        "show_units": True,            # "2700 kg/m3" vs bare "2700"
        "labels": {},                  # property_key -> column header override
    },
    "diff": {
        # Change-highlighting vs a previous version of the export
        # (--diff-against). Only columns that flow into the report are
        # compared (mapped + passthrough), so PDM churn like "Checked Out
        # By" never lights a row green; list report columns here to exclude
        # them from the comparison too.
        "ignore_columns": [],
    },
    "budget": {
        # Spec/RFQ budget rollup (--budget / --dashboard outputs). Parts are
        # flagged into procurement categories by the spec-document reference
        # in the column mapped by columns.specs; a cell may list several
        # separated by spec_separator (the first is the category used for
        # cost rollup). A spec'd node covers its whole subtree; leaf parts
        # under no spec are reported as unassigned.
        "spec_separator": ";",
    },
}


def load_config(path: Path | None) -> dict:
    cfg = json.loads(json.dumps(DEFAULT_CONFIG))  # deep copy
    candidate = path or Path("bomgen.toml")
    if candidate.exists():
        if tomllib is None:
            sys.exit("bomgen: python >=3.11 required for TOML config")
        with open(candidate, "rb") as f:
            user = tomllib.load(f)
        for section, vals in user.items():
            cfg.setdefault(section, {}).update(vals)
        cfg["_config_path"] = str(candidate)
    elif path is not None:
        sys.exit(f"bomgen: config not found: {path}")
    return cfg


# --------------------------------------------------------------------------- model

@dataclass
class BomNode:
    path: str
    depth: int
    qty: int
    filename: str
    raw: dict
    parent: "BomNode | None" = None
    children: list = field(default_factory=list)
    qty_total: int = 1
    part_number: str = ""
    display_name: str = ""
    is_assembly: bool = False
    cots: str = ""
    state: str = ""
    file_url: str = ""
    material_props: dict = field(default_factory=dict)
    specs: list = field(default_factory=list)  # spec refs listed on this row
    spec: str = ""                             # primary (first) spec ref
    diff: str = ""                             # ""|"changed"|"added" vs prev


class ValidationError(Exception):
    pass


def read_csv(path: Path) -> tuple[list[str], list[dict]]:
    last_err = None
    for enc in ("utf-8-sig", "utf-16", "cp1252"):
        try:
            with open(path, encoding=enc, newline="") as f:
                rows = list(csv.DictReader(f))
            if not rows:
                sys.exit(f"bomgen: no data rows in {path}")
            return list(rows[0].keys()), rows
        except (UnicodeDecodeError, UnicodeError) as e:
            last_err = e
    sys.exit(f"bomgen: cannot decode {path}: {last_err}")


def read_xml(path: Path, cfg: dict, warnings: list[str]) -> tuple[list[str], list[dict]]:
    """Ingest a PDM Professional XML export (export-rule / ERP channel).

    Schema-tolerant (see design §2.5 / open item O2): PDM XML wraps data in
    <transactions><transaction><document><configuration><attribute name=.. value=../>.
    Hierarchy may arrive as (a) nested <document> elements under <references>,
    or (b) a dotted-level attribute matching columns.level. Output is
    normalized to CSV-shaped row dicts so build_tree() is shared unchanged.
    """
    import xml.etree.ElementTree as ET  # trusted-vault input; see README note
    col = cfg["columns"]
    try:
        xroot = ET.parse(path).getroot()
    except ET.ParseError as e:
        sys.exit(f"bomgen: XML parse error in {path}: {e}")

    def attrs_of(doc) -> dict:
        d = {}
        for conf in doc.findall("configuration"):
            if conf.get("name") and col["config"] not in d:
                d[col["config"]] = conf.get("name")
            for a in conf.findall("attribute"):
                d[a.get("name", "")] = a.get("value", "")
        # document-level id doubles as filename/number when no explicit attr
        d.setdefault(col["number"], doc.get("id", ""))
        return d

    rows: list[dict] = []
    transactions = xroot.findall(".//transaction") or [xroot]
    top_docs = [d for t in transactions for d in t.findall("document")]
    nested = any(d.find(".//references/document") is not None for d in top_docs)

    if nested:
        counters: dict[str, int] = {}

        def walk(doc, parent_path: str):
            counters[parent_path] = counters.get(parent_path, 0) + 1
            p = (f"{parent_path}.{counters[parent_path]}"
                 if parent_path else str(counters[parent_path]))
            row = attrs_of(doc)
            row[col["level"]] = p
            row.setdefault(col["qty"], doc.get("quantity", "1"))
            rows.append(row)
            refs = doc.find("references")
            if refs is not None:
                for child in refs.findall("document"):
                    # per-reference quantity may sit on the child element
                    walk(child, p)

        for d in top_docs:
            walk(d, "")
        warnings.append("XML: hierarchy from nested <references>; item numbers "
                        "synthesized in document order (O2: verify vs vault export)")
    else:
        for d in top_docs:
            row = attrs_of(d)
            row.setdefault(col["qty"], d.get("quantity", "1"))
            rows.append(row)
        if rows and col["level"] not in rows[0]:
            for i, r in enumerate(rows, 1):
                r[col["level"]] = str(i)
            warnings.append("XML: no level attribute found; treating as flat "
                            "single-level BOM (O2)")

    if not rows:
        sys.exit(f"bomgen: no <document> elements found in {path}")
    header = sorted({k for r in rows for k in r})
    return header, rows


# unresolved SolidWorks property expression, e.g. 'SW-Mass@.SLDPRT': the
# data-card variable is linked to a model property but the evaluated value
# was never cached (file not rebuilt/saved) — the cell is not a number.
SW_PROP_RE = re.compile(r"^SW-\w+@")


def scan_unresolved_props(rows: list[dict], warnings: list[str]) -> None:
    """V7: flag cells that export the raw SW property expression instead of
    its evaluated value. Grouped per column so the banner stays short."""
    hits: dict[str, tuple[int, str]] = {}
    for r in rows:
        for k, v in r.items():
            v = (v or "").strip() if isinstance(v, str) else ""
            if SW_PROP_RE.match(v):
                n, ex = hits.get(k, (0, v))
                hits[k] = (n + 1, ex)
    for k, (n, ex) in sorted(hits.items()):
        warnings.append(
            f"V7: column '{k}': {n} row(s) hold an unresolved SolidWorks "
            f"property expression (e.g. '{ex}'), not a value — rebuild and "
            "save those files in CAD so PDM exports the evaluated number")


def build_tree(header: list[str], rows: list[dict], cfg: dict,
               warnings: list[str]) -> BomNode:
    col = cfg["columns"]
    errors: list[str] = []
    scan_unresolved_props(rows, warnings)

    # V5 required columns
    for key in ("level", "qty", "number"):
        if col[key] not in header:
            errors.append(f"V5: required column '{col[key]}' missing from input")
    if col["cots"] not in header:
        warnings.append(
            f"V5: COTS column '{col['cots']}' not in input; COTS will be blank "
            "(per design D1 — add the variable in PDM or edit the CSV)")
    if errors:
        raise ValidationError("\n".join(errors))

    p = cfg["project"]
    root = BomNode(path="0", depth=0, qty=1, filename="", raw={},
                   is_assembly=True)
    root.part_number = p["assembly_number"]
    root.display_name = p["assembly_name"]
    root.raw = {"Description": p["assembly_description"]}
    index: dict[str, BomNode] = {"0": root}

    prev_depth = 0
    for i, row in enumerate(rows, start=2):  # 1-based incl. header
        path = (row.get(col["level"]) or "").strip()
        if not path:
            warnings.append(f"row {i}: empty Level; row skipped")
            continue
        if path in index:
            # V2/R1: Excel round-trips float-mangle dotted levels ("2.10"->"2.1").
            # Repair iff appending zeros to the last segment yields exactly the
            # next expected sibling number under the same parent.
            repaired = None
            head, _, last = path.rpartition(".")
            pp = head or "0"
            if head and pp in index and last.isdigit():
                expected = len(index[pp].children) + 1
                cand = last
                while len(cand) < 6:
                    cand += "0"
                    if int(cand) == expected:
                        repaired = f"{head}.{cand}"
                        break
                    if int(cand) > expected:
                        break
            if repaired and repaired not in index:
                warnings.append(
                    f"R1: row {i}: duplicate path '{path}' reinterpreted as "
                    f"'{repaired}' (Excel float-mangling; export direct from PDM to avoid)")
                path = repaired
            else:
                errors.append(f"V2: duplicate path '{path}' at row {i}")
                continue
        depth = path.count(".") + 1
        raw_number = row.get(col["number"]) or ""
        filename = raw_number.strip()

        # V3 indent cross-check
        lead = len(raw_number) - len(raw_number.lstrip(" "))
        if raw_number and lead and lead // 2 != depth - 1:
            warnings.append(
                f"V3: row {i} ('{path}'): indent {lead} spaces disagrees with depth {depth}")
        # V6 depth jump
        if depth > prev_depth + 1:
            warnings.append(f"V6: row {i} ('{path}'): depth jumps {prev_depth}->{depth}")
        prev_depth = depth

        # V4 qty
        qty_raw = (row.get(col["qty"]) or "").strip()
        try:
            qty = int(float(qty_raw))
            if qty <= 0:
                raise ValueError
        except ValueError:
            warnings.append(f"V4: row {i} ('{path}'): qty '{qty_raw}' invalid; using 1")
            qty = 1

        parent_path = path.rsplit(".", 1)[0] if "." in path else "0"
        parent = index.get(parent_path)
        if parent is None:
            errors.append(f"V1: row {i} ('{path}'): parent '{parent_path}' not found")
            continue

        node = BomNode(path=path, depth=depth, qty=qty, filename=filename,
                       raw=row, parent=parent)
        parent.children.append(node)
        index[path] = node

    if errors:
        raise ValidationError("\n".join(errors))
    return root


def _found_in_rest(found_in: str, strip: str) -> str:
    """Found In vault path -> URL path tail (backslashes -> "/", each segment
    URL-encoded). The leading drive letter is always dropped. Then:

    - `strip` set: remove that literal prefix (itself drive-tolerant), e.g.
      "D:\\Vault" or just "Vault".
    - `strip` empty (default): also drop the first folder — the vault root —
      matching any `<drive>:\\<Vault>\\…` path, since the URL base already
      ends in the vault name. E.g. "E:\\Obs_PDM\\STP\\ESC\\Flight" ->
      "STP/ESC/Flight".
    """
    def drop_drive(p: str) -> str:
        return re.sub(r"^[A-Za-z]:", "", p.strip().replace("\\", "/")).strip("/")

    fi = drop_drive(found_in or "")
    s = drop_drive(strip or "")
    if s:
        if fi.lower().startswith(s.lower()):
            fi = fi[len(s):].lstrip("/")
    else:
        # no explicit prefix -> drop the vault-root folder automatically
        parts = fi.split("/", 1)
        fi = parts[1] if len(parts) > 1 else ""
    return "/".join(quote(seg, safe="") for seg in fi.split("/") if seg)


def _build_file_url(template: str, found_in: str, strip: str,
                    filename: str) -> str:
    """Substitute {found_in}/{file} in the PDM URL template. Returns "" when
    a placeholder the template uses has no source value (so no dead link)."""
    if not template:
        return ""
    url = template
    if "{found_in}" in url:
        rest = _found_in_rest(found_in, strip)
        if not rest:
            return ""
        url = url.replace("{found_in}", rest)
    if "{file}" in url:
        if not filename:
            return ""
        url = url.replace("{file}", quote(filename, safe=""))
    return url


def derive(root: BomNode, cfg: dict) -> None:
    col, rules, proj = cfg["columns"], cfg["rules"], cfg["project"]
    pn_re = re.compile(rules["part_number_pattern"])
    links = cfg.get("links", {})
    url_tmpl = links.get("file_url_template", "")
    found_in_strip = links.get("found_in_strip", "")
    spec_sep = cfg.get("budget", {}).get("spec_separator", ";")

    def visit(n: BomNode):
        if n.parent is not None:
            n.qty_total = n.qty * n.parent.qty_total
            stem = re.sub(r"\.(sldasm|sldprt|slddrw)$", "", n.filename, flags=re.I)
            n.is_assembly = bool(n.children) or n.filename.lower().endswith(".sldasm")
            m = pn_re.match(stem)
            if m:
                rev = (n.raw.get(col["rev"]) or "").strip()
                dash = rev.zfill(2) if rev.isdigit() else (rev or "00")  # O1
                n.part_number = f"{m.group(1)}-{dash}"
            else:
                n.part_number = stem
            name = (n.raw.get(col["name"]) or "").strip()
            desc = (n.raw.get(col["description"]) or "").strip()
            base = name or desc or stem.replace("_", " ")
            n.display_name = f"{base} ({n.filename})" if n.filename else base
            # PDM viewer link built from the configured template, using the
            # Found In vault path and/or the filename (URL-encoded).
            if url_tmpl:
                n.file_url = _build_file_url(
                    url_tmpl, n.raw.get(col["found_in"]), found_in_strip,
                    n.filename)
            n.cots = (n.raw.get(col["cots"]) or "").strip()
            raw_specs = (n.raw.get(col["specs"]) or "").strip()
            if raw_specs:
                n.specs = [s.strip() for s in raw_specs.split(spec_sep)
                           if s.strip()]
                n.spec = n.specs[0] if n.specs else ""
            if rules["state_from_found_in"]:
                fi = (n.raw.get(col["found_in"]) or "").replace("/", "\\")
                parts = [s for s in fi.split("\\") if s]
                n.state = "\\".join(parts[-2:]) if parts else ""
        for c in n.children:
            visit(c)

    visit(root)


def preorder(root: BomNode) -> list[BomNode]:
    out: list[BomNode] = []

    def walk(n: BomNode):
        out.append(n)
        for c in n.children:
            walk(c)

    walk(root)
    return out


# --------------------------------------------------------------------------- materials

def material_cache_key(name: str) -> str:
    """Normalize a material name for matching (whitespace-collapsed, casefold).

    The shared contract between a materials-database export and this reader:
    a BOM row's Material text matches a DB record iff their keys are equal.
    """
    return re.sub(r"\s+", " ", (name or "").strip()).casefold()


def _material_index(docs: list) -> dict[str, dict]:
    """Index a raw materials-database export (the /export/raw-json dump: a
    list of material documents) by normalized name, each synonym aliased to
    the same Properties map. Later duplicates lose to earlier ones."""
    index: dict[str, dict] = {}
    for d in docs:
        if not isinstance(d, dict) or d.get("is_deleted"):
            continue
        name = d.get("Material")
        if not name:
            continue
        props = d.get("Properties") if isinstance(d.get("Properties"), dict) else {}
        for alias in [name, *(d.get("Material_Synonyms") or [])]:
            key = material_cache_key(alias)
            if key and key not in index:
                index[key] = props
    return index


def enrich_materials(root: BomNode, cfg: dict, warnings: list[str]) -> None:
    """Populate each node's material_props from a committed materials-database
    export (config `[materials]`). Local-file only; no network. Purely
    additive and a no-op unless `[materials].enabled`. Cache miss = blank
    properties + a grouped warning, never an abort (cf. COTS / D1)."""
    mat = cfg.get("materials", {})
    if not mat.get("enabled"):
        return
    wanted = list(mat.get("properties", []))
    show_units = mat.get("show_units", True)
    col = cfg["columns"]
    cache_file = mat.get("cache_file", "")
    path = Path(cache_file) if cache_file else None

    index: dict[str, dict] = {}
    if path is None or not path.exists():
        warnings.append(
            f"V8: materials enrichment enabled but cache file "
            f"'{cache_file}' not found; material properties left blank")
    else:
        try:
            docs = json.loads(path.read_text(encoding="utf-8"))
            if isinstance(docs, dict):  # tolerate {"materials": [...]} wrappers
                docs = docs.get("materials") or docs.get("data") or [docs]
            index = _material_index(docs)
        except (json.JSONDecodeError, OSError) as e:
            warnings.append(
                f"V8: materials cache '{path}' unreadable ({e}); "
                "material properties left blank")

    misses: dict[str, int] = {}
    for n in preorder(root):
        if n.parent is None:
            continue
        raw_name = (n.raw.get(col["material"]) or "").strip()
        if not raw_name:
            continue
        entry = index.get(material_cache_key(raw_name)) if index else None
        if entry is None:
            n.material_props = {p: "" for p in wanted}
            if index:  # only a "miss" if we actually have a database to miss in
                misses[raw_name] = misses.get(raw_name, 0) + 1
            continue
        rendered = {}
        for p in wanted:
            prop = entry.get(p)
            if isinstance(prop, dict) and prop.get("value") is not None:
                unit = prop.get("unit") or ""
                rendered[p] = (f"{prop['value']} {unit}".strip()
                               if show_units and unit else str(prop["value"]))
            else:
                rendered[p] = ""
        n.material_props = rendered

    if misses:
        total = sum(misses.values())
        examples = ", ".join(sorted(misses)[:3])
        warnings.append(
            f"V9: {total} row(s) reference {len(misses)} material(s) not found "
            f"in the materials database (e.g. {examples}); properties left blank "
            "— re-export the materials JSON or check the Material names")


# --------------------------------------------------------------------------- diff

def _row_signature(n: BomNode, cfg: dict) -> dict:
    """The comparable content of a row: only columns that flow into the
    report (mapped + passthrough), minus [diff].ignore_columns — so PDM
    churn in unmapped metadata never counts as a change. Level/position is
    deliberately excluded: a move is not a content change."""
    col = cfg["columns"]
    ignore = set(cfg.get("diff", {}).get("ignore_columns", []))
    keys = [col[k] for k in ("qty", "name", "config", "rev", "description",
                              "found_in", "cots", "material", "specs")
            if col.get(k)]
    keys += list(col.get("passthrough", []))
    return {k: (n.raw.get(k) or "").strip() for k in keys if k not in ignore}


def _diff_index(root: BomNode) -> dict:
    """(parent filename, filename, occurrence#) -> node, for identity
    matching that survives item renumbering (never keyed on Level path)."""
    idx: dict = {}
    counts: dict = {}
    def walk(n: BomNode):
        for c in n.children:
            base = (n.filename, c.filename)
            occ = counts.get(base, 0)
            counts[base] = occ + 1
            idx[(n.filename, c.filename, occ)] = c
            walk(c)
    walk(root)
    return idx


def apply_diff(root: BomNode, prev_path: Path, cfg: dict,
               warnings: list[str]) -> None:
    """Mark each node changed/added vs a previous version of the export
    (n.diff), and summarize removals in the warning banner. The previous
    file is parsed with a throwaway warnings list — its own parse issues
    belong to its own build, not this one. Unreadable/invalid previous
    file -> DIFF warning and no highlighting, never an abort."""
    try:
        throwaway: list[str] = []
        if prev_path.suffix.lower() == ".xml":
            p_header, p_rows = read_xml(prev_path, cfg, throwaway)
        else:
            p_header, p_rows = read_csv(prev_path)
        prev_root = build_tree(p_header, p_rows, cfg, throwaway)
        derive(prev_root, cfg)
    except (SystemExit, ValidationError, OSError) as e:
        warnings.append(f"DIFF: previous version '{prev_path}' could not be "
                        f"parsed ({e}); change highlighting skipped")
        return

    prev_idx = _diff_index(prev_root)
    consumed: set = set()
    changed = added = 0

    # fallback pool for moved parts: filename -> [keys] not yet matched
    by_file: dict = {}
    for k in prev_idx:
        by_file.setdefault(k[1], []).append(k)

    def match(key):
        if key in prev_idx and key not in consumed:
            consumed.add(key)
            return prev_idx[key]
        # moved: same filename under a different parent / occurrence
        for k in by_file.get(key[1], []):
            if k not in consumed:
                consumed.add(k)
                return prev_idx[k]
        return None

    cur_idx = _diff_index(root)
    for key, n in cur_idx.items():
        prev_n = match(key)
        if prev_n is None:
            n.diff = "added"
            added += 1
        elif _row_signature(n, cfg) != _row_signature(prev_n, cfg):
            n.diff = "changed"
            changed += 1

    cur_files = {n.filename for n in preorder(root) if n.filename}
    removed = sorted({k[1] for k in prev_idx
                      if k not in consumed and k[1] not in cur_files})
    if changed or added or removed:
        msg = (f"DIFF: vs previous version: {changed} row(s) changed, "
               f"{added} added (highlighted green)")
        if removed:
            msg += (f"; {len(removed)} part(s) removed "
                    f"(e.g. {', '.join(removed[:3])})")
        warnings.append(msg)


# --------------------------------------------------------------------------- budget

UNASSIGNED = "(unassigned)"


def budget_rollup(root: BomNode, cfg: dict, warnings: list[str],
                  header: list[str] | None = None) -> dict:
    """Group the BOM into RFQ/procurement categories keyed by spec document.

    Semantics (see design §5.4): a node carrying a spec reference is a budget
    line item, and its whole subtree is considered covered by that line (a
    quoted weldment includes its pieces). Leaf parts under no spec'd ancestor
    and with no spec of their own land in UNASSIGNED so budget gaps are
    visible, not silent. Identical parts (same filename) under the same spec
    merge into one line with summed total quantity.
    Warnings: V10 specs column missing / parts unassigned, V11 multi-spec
    cells (only the first is used for cost rollup), V12 spec'd items nested
    inside another spec'd subtree (possible double-count).
    """
    col = cfg["columns"]
    spec_url_tmpl = cfg.get("links", {}).get("spec_url_template", "")
    groups: dict[str, dict] = {}
    multi: dict[str, int] = {}
    nested: dict[str, int] = {}
    unassigned_parts: dict[str, int] = {}

    def add_line(spec: str, n: BomNode) -> None:
        g = groups.setdefault(spec, {"name": spec, "url": "", "lines": {}})
        key = n.filename or n.part_number or n.path
        line = g["lines"].get(key)
        if line is None:
            g["lines"][key] = {
                "pn": n.part_number, "name": n.display_name,
                "file": n.filename, "fileUrl": n.file_url,
                "cots": n.cots,
                "config": (n.raw.get(col["config"]) or "").strip(),
                "asm": n.is_assembly,
                "qty": n.qty_total, "occ": 1, "diff": n.diff,
            }
        else:
            line["qty"] += n.qty_total
            line["occ"] += 1
            if n.diff and not line["diff"]:
                line["diff"] = n.diff

    def visit(n: BomNode, covering: str) -> None:
        if n.parent is not None:
            if n.spec:
                if len(n.specs) > 1:
                    multi[n.filename or n.path] = len(n.specs)
                if covering:
                    nested[n.filename or n.path] = 1
                add_line(n.spec, n)
                covering = n.spec
            elif not covering and not n.children:
                add_line(UNASSIGNED, n)
                unassigned_parts[n.filename or n.path] = 1
        for c in n.children:
            visit(c, covering)

    visit(root, "")

    if header is not None and col["specs"] not in header:
        warnings.append(
            f"V10: specs column '{col['specs']}' not in input; every part is "
            "unassigned — add the column in the PDM export or map "
            "columns.specs in the config")
    elif unassigned_parts:
        examples = ", ".join(sorted(unassigned_parts)[:3])
        warnings.append(
            f"V10: {len(unassigned_parts)} part(s) not covered by any spec "
            f"(e.g. {examples}) — listed under {UNASSIGNED} in the budget")
    if multi:
        examples = ", ".join(sorted(multi)[:3])
        warnings.append(
            f"V11: {len(multi)} row(s) list multiple specs (e.g. {examples}); "
            "only the first is used for the cost rollup")
    if nested:
        examples = ", ".join(sorted(nested)[:3])
        warnings.append(
            f"V12: {len(nested)} spec'd item(s) nested inside another spec'd "
            f"subtree (e.g. {examples}) — check for double-counted cost")

    specs_out = []
    for name in sorted(k for k in groups if k != UNASSIGNED):
        g = groups[name]
        lines = list(g["lines"].values())
        if spec_url_tmpl:
            g["url"] = spec_url_tmpl.replace("{spec}", quote(name, safe=""))
        specs_out.append({"name": name, "url": g["url"], "lines": lines,
                          "totalQty": sum(l["qty"] for l in lines)})
    ug = groups.get(UNASSIGNED, {"lines": {}})
    ulines = list(ug["lines"].values())
    return {
        "specs": specs_out,
        "unassigned": {"name": UNASSIGNED, "url": "", "lines": ulines,
                        "totalQty": sum(l["qty"] for l in ulines)},
        "counts": {
            "specs": len(specs_out),
            "lines": sum(len(s["lines"]) for s in specs_out),
            "unassignedLines": len(ulines),
            "totalQty": sum(s["totalQty"] for s in specs_out)
                        + sum(l["qty"] for l in ulines),
            "changed": sum(1 for s in specs_out for l in s["lines"]
                           if l.get("diff"))
                       + sum(1 for l in ulines if l.get("diff")),
        },
    }


DIFF_GREEN = "E2EFDA"  # light green fill for rows changed/added vs previous
HIST_YELLOW = "FFF3C4"


def write_budget_xlsx(rollup: dict, cfg: dict, src_name: str, out: Path,
                      warnings: list[str] | None = None,
                      source_rev: str = "", historical: str = "",
                      source_url: str = "") -> None:
    """Budget workbook: sheet "Budget" = per-spec rollup driven by live
    SUMIF/COUNTIF formulas over sheet "Parts", the costing template where
    unit WAG / ROM / Quote get typed in (shaded input columns). Ext costs and
    the Best chain (Quote > ROM > WAG) are formulas, so the workbook keeps
    rolling up as estimates mature — the point of the template. (This is a
    working costing sheet, unlike the static BOM report; formulas are
    deliberate here.)"""
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    fontname = cfg["output"]["excel_font"]
    proj = cfg["project"]
    F = lambda size, bold=False: Font(name=fontname, size=size, bold=bold)
    input_fill = PatternFill("solid", fgColor="E8F0E8")
    head_border = Border(bottom=Side(style="medium"))
    money = "$#,##0.00"

    wb = openpyxl.Workbook()

    # ---------------- Parts sheet (the costing template)
    ps = wb.create_sheet("Parts")
    p_headers = ["Spec", "Part Number", "Part Name", "Config", "COTS",
                 "Qty (Total)", "Unit WAG", "Unit ROM", "Unit Quote",
                 "Unit Best", "Ext WAG", "Ext ROM", "Ext Quote", "Ext Best"]
    for j, h in enumerate(p_headers, 1):
        c = ps.cell(row=1, column=j, value=h)
        c.font, c.border = F(11, True), head_border
    r = 2
    all_groups = rollup["specs"] + ([rollup["unassigned"]]
                                    if rollup["unassigned"]["lines"] else [])
    for g in all_groups:
        for line in g["lines"]:
            ps.cell(row=r, column=1, value=g["name"])
            ps.cell(row=r, column=2, value=line["pn"])
            ps.cell(row=r, column=3, value=line["name"])
            ps.cell(row=r, column=4, value=line["config"])
            ps.cell(row=r, column=5, value=line["cots"])
            ps.cell(row=r, column=6, value=line["qty"])
            for cidx in (7, 8, 9):  # unit cost inputs
                cell = ps.cell(row=r, column=cidx)
                cell.fill, cell.number_format = input_fill, money
            ps.cell(row=r, column=10,
                    value=f'=IF(I{r}<>"",I{r},IF(H{r}<>"",H{r},'
                          f'IF(G{r}<>"",G{r},"")))')
            ps.cell(row=r, column=11, value=f'=IF(G{r}<>"",$F{r}*G{r},"")')
            ps.cell(row=r, column=12, value=f'=IF(H{r}<>"",$F{r}*H{r},"")')
            ps.cell(row=r, column=13, value=f'=IF(I{r}<>"",$F{r}*I{r},"")')
            ps.cell(row=r, column=14, value=f'=IF(J{r}<>"",$F{r}*J{r},"")')
            for cidx in range(10, 15):
                ps.cell(row=r, column=cidx).number_format = money
            for cidx in range(1, 15):
                ps.cell(row=r, column=cidx).font = F(11)
            if line.get("diff"):  # changed/added vs previous version
                green = PatternFill("solid", fgColor=DIFF_GREEN)
                for cidx in (1, 2, 3, 4, 5, 6):  # data cells, not cost inputs
                    ps.cell(row=r, column=cidx).fill = green
            r += 1
    p_last = r - 1
    for j, w in enumerate([22, 24, 52, 14, 7, 11, 10, 10, 10, 10,
                            12, 12, 12, 12], 1):
        ps.column_dimensions[get_column_letter(j)].width = w
    ps.freeze_panes = "A2"
    ps.auto_filter.ref = f"A1:N{max(p_last, 1)}"

    # ---------------- Budget sheet (rollup by spec, formula-linked)
    bs = wb.active
    bs.title = "Budget"
    t = bs.cell(row=1, column=1, value="Hardware Budget by Specification")
    t.font = F(16, True)
    meta = f"{proj['title_line']} · from {src_name}"
    if source_rev:
        meta += f" · rev {source_rev}"
    meta += f" · generated {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    meta_cell = bs.cell(row=2, column=1, value=meta)
    meta_cell.font = F(9)
    if source_url:
        meta_cell.hyperlink = source_url
        meta_cell.font = Font(name=fontname, size=9, color="0563C1",
                              underline="single")
    bs.cell(row=3, column=1,
            value="Enter unit costs on the Parts sheet (shaded WAG/ROM/Quote "
                  "columns); every figure here rolls up live. Best = Quote, "
                  "else ROM, else WAG.").font = F(9)
    if historical:
        h = bs.cell(row=5, column=1,
                    value=f"⚠ HISTORICAL VERSION — {historical}")
        h.font = Font(name=fontname, size=11, bold=True, color="7A4E00")
        h.fill = PatternFill("solid", fgColor=HIST_YELLOW)
        bs.merge_cells(start_row=5, start_column=1, end_row=5, end_column=8)
    if warnings:
        wcell = bs.cell(row=4, column=1, value="⚠ " + " · ".join(
            banner_lines(warnings)))
        wcell.font = Font(name=fontname, size=9, color="7A4E00")
        wcell.fill = PatternFill("solid", fgColor="FFF3C4")
        bs.merge_cells(start_row=4, start_column=1, end_row=4, end_column=8)
        wcell.alignment = Alignment(wrap_text=True, vertical="top")
        bs.row_dimensions[4].height = max(15, 12 * len(banner_lines(warnings)))

    b_headers = ["Spec", "Line Items", "Total Qty", "Ext WAG", "Ext ROM",
                 "Ext Quote", "Ext Best", "Spec Document"]
    hrow = 6
    for j, h in enumerate(b_headers, 1):
        c = bs.cell(row=hrow, column=j, value=h)
        c.font, c.border = F(11, True), head_border
    r = hrow + 1
    names = [g["name"] for g in all_groups]
    for name in names:
        bs.cell(row=r, column=1, value=name).font = F(11, True)
        bs.cell(row=r, column=2, value=f'=COUNTIF(Parts!$A:$A,A{r})')
        bs.cell(row=r, column=3, value=f'=SUMIF(Parts!$A:$A,A{r},Parts!$F:$F)')
        bs.cell(row=r, column=4, value=f'=SUMIF(Parts!$A:$A,A{r},Parts!$K:$K)')
        bs.cell(row=r, column=5, value=f'=SUMIF(Parts!$A:$A,A{r},Parts!$L:$L)')
        bs.cell(row=r, column=6, value=f'=SUMIF(Parts!$A:$A,A{r},Parts!$M:$M)')
        bs.cell(row=r, column=7, value=f'=SUMIF(Parts!$A:$A,A{r},Parts!$N:$N)')
        for j in range(4, 8):
            bs.cell(row=r, column=j).number_format = money
        for j in range(2, 8):
            bs.cell(row=r, column=j).font = F(11)
        url = next((g["url"] for g in all_groups if g["name"] == name), "")
        if url:
            c = bs.cell(row=r, column=8, value=url)
            c.hyperlink, c.font = url, Font(name=fontname, size=10,
                                            color="0563C1",
                                            underline="single")
        r += 1
    tr = bs.cell(row=r, column=1, value="TOTAL")
    tr.font = F(12, True)
    first_data = hrow + 1
    for j, letter in ((2, "B"), (3, "C"), (4, "D"), (5, "E"), (6, "F"),
                      (7, "G")):
        c = bs.cell(row=r, column=j,
                    value=f"=SUM({letter}{first_data}:{letter}{r - 1})")
        c.font = F(12, True)
        if j >= 4:
            c.number_format = money
    for j, w in enumerate([26, 11, 10, 13, 13, 13, 13, 40], 1):
        bs.column_dimensions[get_column_letter(j)].width = w
    bs.freeze_panes = f"A{hrow + 1}"

    wb.save(out)


# --------------------------------------------------------------------------- excel

BANNER_MAX = 8  # warnings shown in the data-quality banner before "+N more"


def banner_lines(warnings: list[str]) -> list[str]:
    shown = warnings[:BANNER_MAX]
    if len(warnings) > BANNER_MAX:
        shown.append(f"…and {len(warnings) - BANNER_MAX} more (see generator stderr)")
    return shown


def write_excel(root: BomNode, cfg: dict, out: Path,
                warnings: list[str] | None = None, source_rev: str = "",
                historical: str = "", src_name: str = "",
                source_url: str = "") -> None:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    proj, colcfg, rules = cfg["project"], cfg["columns"], cfg["rules"]
    fontname = cfg["output"]["excel_font"]
    nodes = preorder(root)
    max_depth = max(n.depth for n in nodes)
    n_mark = max(rules["min_marker_columns"], max_depth)
    passthrough = list(colcfg.get("passthrough", []))
    matcfg = cfg.get("materials", {})
    mat_props = list(matcfg.get("properties", [])) if matcfg.get("enabled") else []
    mat_headers = [matcfg.get("labels", {}).get(p, p.replace("_", " "))
                   for p in mat_props]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BOM"
    ws.sheet_properties.outlinePr.summaryBelow = False

    F = lambda size, bold=False: Font(name=fontname, size=size, bold=bold)
    medium_bottom = Border(bottom=Side(style="medium"))

    first = 2  # column B, matching template
    def L(idx):  # 0-based logical index -> letter
        return get_column_letter(first + idx)

    headers = ["Level", "Abbreviated Part Number", "Part Name", "Description",
               "COTS", "Qty (Assembly)", "Qty (Total)"] + passthrough + mat_headers
    total_cols = n_mark + len(headers)

    # ---- data-quality banner (rows 2-5 are blank in the template, so the
    # banner never shifts the title block / header rows the team knows)
    if warnings:
        lines = ["⚠ DATA QUALITY WARNINGS — verify before use:"] + [
            f"• {w}" for w in banner_lines(warnings)]
        yellow = PatternFill("solid", fgColor="FFF3C4")
        edge = Side(style="thin", color="B98900")
        box = Border(left=edge, right=edge, top=edge, bottom=edge)
        for row in range(2, 6):
            for i in range(total_cols):
                cell = ws[f"{L(i)}{row}"]
                cell.fill, cell.border = yellow, box
        c = ws["B2"]
        c.value = "\n".join(lines)
        c.font = Font(name=fontname, size=11, bold=False, color="7A4E00")
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(f"B2:{L(total_cols - 1)}5")
        est = sum(len(l) // 160 + 1 for l in lines)  # rough wrap estimate
        for row in range(2, 6):
            ws.row_dimensions[row].height = max(15, est * 15 / 4 + 2)

    # ---- historical marker (row 1 is blank in the template)
    if historical:
        h = ws["B1"]
        h.value = f"⚠ HISTORICAL VERSION — {historical}"
        h.font = Font(name=fontname, size=12, bold=True, color="7A4E00")
        h.fill = PatternFill("solid", fgColor=HIST_YELLOW)
        ws.merge_cells(f"B1:{L(total_cols - 1)}1")

    # ---- title block
    def put(row, text, size=14, bold=True):
        c = ws[f"B{row}"]
        c.value, c.font = text, F(size, bold)
        ws.merge_cells(f"B{row}:{L(n_mark)}{row}")

    put(6, "Bill of Materials (BOM)", 18)
    put(7, proj["title_line"])
    if src_name or source_rev:
        # row 8 is a blank spacer in the template; safe to use for
        # provenance without disturbing the documented title-block rows
        if src_name and source_rev:
            row8_text = f"Source: {src_name} · rev {source_rev}"
        elif src_name:
            row8_text = f"Source: {src_name}"
        else:
            row8_text = f"Source revision: {source_rev}"
        put(8, row8_text, size=10, bold=False)
        if source_url:
            c = ws["B8"]
            c.hyperlink = source_url
            c.font = Font(name=fontname, size=10, bold=False,
                          color="0563C1", underline="single")
    put(9, proj["project_box"])
    put(10, f"System Name: {proj['system_name']}")
    put(11, f"Assembly Name {proj['assembly_number']}")
    put(12, proj["contact_name"] or "Contact Name")
    put(13, proj["contact_info"] or "Contact Info")

    # ---- header band (row 16) + header row (17)
    b = ws["B16"]
    b.value, b.font, b.border = "Assembly Level", F(18, True), medium_bottom
    ws.merge_cells(f"B16:{L(n_mark - 1)}16")
    qa_idx, qt_idx = n_mark + 5, n_mark + 6
    q = ws[f"{L(qa_idx)}16"]
    q.value, q.font, q.border = "Quantity Required", F(14, True), medium_bottom
    ws.merge_cells(f"{L(qa_idx)}16:{L(qt_idx)}16")
    for i in range(total_cols):
        ws[f"{L(i)}16"].border = medium_bottom

    for i in range(n_mark):
        c = ws[f"{L(i)}17"]
        c.value, c.font, c.border = str(i + 1), F(14, True), medium_bottom
        c.alignment = Alignment(horizontal="center")
    for j, h in enumerate(headers):
        c = ws[f"{L(n_mark + j)}17"]
        c.value, c.font, c.border = h, F(14, True), medium_bottom

    # ---- data rows
    r = 18
    for n in nodes:
        desc = (n.raw.get(colcfg["description"]) or "").strip()
        if n.depth == 0:
            for i in range(n_mark):  # root: X across the band (per template example)
                ws[f"{L(i)}{r}"] = "X"
        else:
            ws[f"{L(n.depth - 1)}{r}"] = "X"
        vals = [n.depth, n.part_number, n.display_name, desc, n.cots,
                n.qty, n.qty_total] + [
                    (n.raw.get(p) or "").strip() for p in passthrough] + [
                    n.material_props.get(p, "") for p in mat_props]
        diff_fill = (PatternFill("solid", fgColor=DIFF_GREEN)
                     if n.diff else None)
        for j, v in enumerate(vals):
            c = ws[f"{L(n_mark + j)}{r}"]
            c.value = v
            c.font = F(14 if n.depth == 0 else 11, n.is_assembly)
            if diff_fill:
                c.fill = diff_fill
        if n.file_url:
            # Part Name is logical column index 2; make the whole cell a
            # hyperlink to the PDM viewer (xlsx can't link a substring).
            c = ws[f"{L(n_mark + 2)}{r}"]
            c.hyperlink = n.file_url
            c.font = Font(name=fontname, size=(14 if n.depth == 0 else 11),
                          bold=n.is_assembly, color="0563C1", underline="single")
        for i in range(n_mark):
            ws[f"{L(i)}{r}"].font = F(14 if n.depth == 0 else 11, True)
            ws[f"{L(i)}{r}"].alignment = Alignment(horizontal="center")
        if n.depth > 0:
            ws.row_dimensions[r].outline_level = min(n.depth, 7)
        r += 1

    # ---- widths, freeze, autofilter
    for i in range(n_mark):
        ws.column_dimensions[L(i)].width = 5.3
    for j, w in enumerate([9, 26, 62, 55, 8, 15, 12]
                          + [14] * len(passthrough) + [14] * len(mat_props)):
        ws.column_dimensions[L(n_mark + j)].width = w
    ws.freeze_panes = f"B18"
    ws.auto_filter.ref = f"{L(n_mark)}17:{L(total_cols - 1)}{r - 1}"

    wb.save(out)


# --------------------------------------------------------------------------- html

def write_html(root: BomNode, cfg: dict, src_name: str, out: Path,
               warnings: list[str], xlsx_href: str = "",
               source_rev: str = "", historical: str = "",
               current_href: str = "../../index.html",
               dashboard_href: str = "",
               source_url: str = "") -> None:
    """xlsx_href: URL/relative path to the sibling .xlsx (empty -> the
    download button is removed client-side). When publishing to GitHub or
    GitLab Pages the .xlsx is deployed next to the HTML, so a bare filename
    works on both (see PAGES_SETUP.md).

    source_rev: opaque provenance string (e.g. the input CSV's last git
    commit hash in the repo that owns it) rendered next to the source
    filename. bomgen itself has no git dependency — callers compute this
    (see template-repo/scripts/build_pages.sh) and pass it in verbatim.

    source_url: URL pointing to the source BOM file in its repository (e.g.
    a GitHub blob URL). When provided the source filename is rendered as a
    hyperlink; empty -> plain ``<code>`` text (no dead link)."""
    proj = cfg["project"]
    passthrough = list(cfg["columns"].get("passthrough", []))
    matcfg = cfg.get("materials", {})
    mat_props = list(matcfg.get("properties", [])) if matcfg.get("enabled") else []
    mat_headers = [matcfg.get("labels", {}).get(p, p.replace("_", " "))
                   for p in mat_props]
    extra_cols = passthrough + mat_headers
    nodes = preorder(root)
    data = [{
        "path": n.path, "depth": n.depth, "pn": n.part_number,
        "name": n.display_name,
        "desc": (n.raw.get(cfg["columns"]["description"]) or "").strip(),
        "cots": n.cots, "qty": n.qty, "qtyTotal": n.qty_total,
        "state": n.state, "asm": n.is_assembly,
        "file": n.filename, "fileUrl": n.file_url, "diff": n.diff,
        "extra": [(n.raw.get(p) or "").strip() for p in passthrough]
                 + [n.material_props.get(p, "") for p in mat_props],
    } for n in nodes]

    tmpl = Path(__file__).with_name("template.html").read_text(encoding="utf-8")
    page = (tmpl
            .replace("/*__DATA__*/", json.dumps(
                {"rows": data, "extraCols": extra_cols}, ensure_ascii=False))
            .replace("__TITLE__", html.escape(proj["title_line"]))
            .replace("__SYSTEM__", html.escape(proj["system_name"]))
            .replace("__ASSY__", html.escape(proj["assembly_number"]))
            .replace("__CONTACT__", html.escape(
                " · ".join(x for x in (proj["contact_name"], proj["contact_info"]) if x)))
            .replace("__GENERATED__", datetime.now().strftime("%Y-%m-%d %H:%M"))
            .replace("__SOURCE__", _source_link_html(src_name, source_url))
            .replace("__SOURCE_REV__",
                     f" · rev <code>{html.escape(source_rev)}</code>" if source_rev else "")
            .replace("__XLSX_HREF__", html.escape(xlsx_href, quote=True))
            .replace("__DASHBOARD_HREF__", html.escape(dashboard_href, quote=True))
            .replace("__HISTORICAL__", _historical_html(historical, current_href))
            .replace("__PROVENANCE_DETAILS__",
                     _provenance_details_html(src_name, source_url, source_rev, cfg))
            .replace("__WARNBOX__", _warnbox_html(warnings)))
    out.write_text(page, encoding="utf-8")


def _historical_html(label: str, current_href: str) -> str:
    """Yellow page chrome for pages built from a non-current (tagged)
    version: tint the background and pin an alert bar linking to current."""
    if not label:
        return ""
    return (
        "<style>body{background:#fdf3d0}#controls{background:#fdf3d0}"
        "header{border-bottom-color:#b98900}</style>"
        '<div id="histbar" style="background:#f5d76e;color:#4a3a00;'
        "padding:8px 22px;font-weight:600;border-bottom:2px solid #b98900\">"
        f"&#9888; Historical version {html.escape(label)} — "
        f'<a href="{html.escape(current_href, quote=True)}" '
        'style="color:#4a3a00">view current</a></div>')


def _warnbox_html(warnings: list[str]) -> str:
    if not warnings:
        return ""
    items = "".join(f"<li>{html.escape(w)}</li>"
                    for w in banner_lines(warnings))
    return ('<div id="warnbox"><strong>&#9888; Data quality warnings '
            f"— verify before use:</strong><ul>{items}</ul></div>")


def _source_link_html(src_name: str, source_url: str) -> str:
    """Return ``<code>src_name</code>`` or a hyperlinked variant.

    Follows the same "no dead link" rule used by xlsx_href / dashboard_href:
    empty source_url -> plain ``<code>``, no ``<a>``."""
    code = f'<code>{html.escape(src_name)}</code>'
    if source_url:
        return f'<a href="{html.escape(source_url, quote=True)}">{code}</a>'
    return code


def _provenance_details_html(src_name: str, source_url: str,
                              source_rev: str, cfg: dict) -> str:
    """Collapsible ``<details>`` provenance block for the HTML header.

    Shows the full source URL (when provided), source revision, bomgen
    version, and the config file path (when a custom config was used).
    Returns an empty string when there is nothing extra to show."""
    items = []
    if source_url:
        items.append(
            f'Source: <a href="{html.escape(source_url, quote=True)}">'
            f'{html.escape(src_name)}</a>')
    if source_rev:
        items.append(f'Rev: <code>{html.escape(source_rev)}</code>')
    items.append(f'bomgen {html.escape(__version__)}')
    config_path = cfg.get("_config_path")
    if config_path:
        items.append(f'Config: <code>{html.escape(config_path)}</code>')
    inner = ' · '.join(items)
    return (
        '<details class="prov"><summary>Provenance</summary>'
        f'<div class="prov-body">{inner}</div></details>'
    )


def write_dashboard(rollup: dict, cfg: dict, src_name: str, out: Path,
                    warnings: list[str], budget_href: str = "",
                    bom_href: str = "", source_rev: str = "",
                    historical: str = "",
                    current_href: str = "../../dashboard.html",
                    source_url: str = "") -> None:
    """Spec/RFQ budget dashboard: one self-contained page (same packaging as
    the BOM report) rolling the BOM up by spec document — stat tiles, a
    collapsible group per spec with its line items, filtering, and a download
    button for the budget .xlsx twin (relative sibling href, same rule as the
    BOM page's Excel button). Costs live in the workbook, not here — the
    dashboard shows the structural rollup the costs hang off."""
    proj = cfg["project"]
    tmpl = Path(__file__).with_name("dashboard.html").read_text(encoding="utf-8")
    page = (tmpl
            .replace("/*__DATA__*/", json.dumps(rollup, ensure_ascii=False))
            .replace("__TITLE__", html.escape(proj["title_line"]))
            .replace("__SYSTEM__", html.escape(proj["system_name"]))
            .replace("__ASSY__", html.escape(proj["assembly_number"]))
            .replace("__CONTACT__", html.escape(
                " · ".join(x for x in (proj["contact_name"], proj["contact_info"]) if x)))
            .replace("__GENERATED__", datetime.now().strftime("%Y-%m-%d %H:%M"))
            .replace("__SOURCE__", _source_link_html(src_name, source_url))
            .replace("__SOURCE_REV__",
                     f" · rev <code>{html.escape(source_rev)}</code>" if source_rev else "")
            .replace("__BUDGET_XLSX_HREF__", html.escape(budget_href, quote=True))
            .replace("__BOM_HREF__", html.escape(bom_href, quote=True))
            .replace("__HISTORICAL__", _historical_html(historical, current_href))
            .replace("__PROVENANCE_DETAILS__",
                     _provenance_details_html(src_name, source_url, source_rev, cfg))
            .replace("__WARNBOX__", _warnbox_html(warnings)))
    out.write_text(page, encoding="utf-8")


# --------------------------------------------------------------------------- cli

def main(argv=None) -> int:
    ap = argparse.ArgumentParser(prog="bomgen", description=__doc__)
    ap.add_argument("input", type=Path)
    ap.add_argument("-c", "--config", type=Path)
    ap.add_argument("--xlsx", type=Path, nargs="?", const=True, default=None)
    ap.add_argument("--html", type=Path, nargs="?", const=True, default=None)
    ap.add_argument("--both", action="store_true")
    ap.add_argument("--xlsx-url", default=None, metavar="URL",
                    help="href for the HTML download button (default: the "
                         ".xlsx filename when both outputs land in the same "
                         "directory, as on a Pages deploy; omitted otherwise)")
    ap.add_argument("--source-rev", default="", metavar="REV",
                    help="opaque provenance string embedded in both reports "
                         "(e.g. the input file's last git commit hash: "
                         "$(git log -1 --format=%%h -- INPUT)); blank if omitted")
    ap.add_argument("--source-url", default="", metavar="URL",
                    help="URL of the source BOM file (e.g. a GitHub blob URL "
                         "like https://github.com/owner/repo/blob/<sha>/path.csv) "
                         "that turns the source filename into a hyperlink in HTML "
                         "pages and xlsx workbooks; empty -> plain text (no dead link)")
    ap.add_argument("--materials-cache", type=Path, default=None, metavar="PATH",
                    help="raw materials-database export (/export/raw-json) to "
                         "enrich rows from; overrides [materials].cache_file and "
                         "implies enabled=true for this run")
    ap.add_argument("--budget", type=Path, nargs="?", const=True, default=None,
                    help="write the spec/RFQ budget workbook (rollup sheet + "
                         "WAG/ROM/Quote costing template with live formulas)")
    ap.add_argument("--dashboard", type=Path, nargs="?", const=True, default=None,
                    help="write the spec/RFQ budget dashboard page (rollup of "
                         "parts grouped by spec document)")
    ap.add_argument("--diff-against", type=Path, default=None, metavar="PREV",
                    help="previous version of the input (extracted by the "
                         "caller, e.g. via git show) — rows changed/added "
                         "since it are highlighted green in every output; "
                         "removed parts are summarized in the banner")
    ap.add_argument("--historical", default="", metavar="LABEL",
                    help="mark the generated pages as a historical version "
                         "(yellow chrome + banner linking to current); LABEL "
                         "is typically the git tag, e.g. 'v0.3 (2026-07-01)'")
    ap.add_argument("-o", "--outdir", type=Path, default=Path("."))
    ap.add_argument("--quiet", action="store_true")
    a = ap.parse_args(argv)

    cfg = load_config(a.config)
    warnings: list[str] = []
    if a.input.suffix.lower() == ".xml":
        header, rows = read_xml(a.input, cfg, warnings)
    else:
        header, rows = read_csv(a.input)
    try:
        root = build_tree(header, rows, cfg, warnings)
    except ValidationError as e:
        print(f"bomgen: validation errors:\n{e}", file=sys.stderr)
        return 1
    derive(root, cfg)
    if a.diff_against is not None:
        apply_diff(root, a.diff_against, cfg, warnings)
    if a.materials_cache is not None:
        cfg["materials"]["enabled"] = True
        cfg["materials"]["cache_file"] = str(a.materials_cache)
    enrich_materials(root, cfg, warnings)
    rollup = (budget_rollup(root, cfg, warnings, header)
              if (a.budget or a.dashboard) else None)
    if rollup is not None:
        rollup["diffed"] = a.diff_against is not None

    if not a.quiet:
        for w in warnings:
            print(f"warning: {w}", file=sys.stderr)

    a.outdir.mkdir(parents=True, exist_ok=True)
    stem = a.input.stem
    did = False
    xlsx_out: Path | None = None
    html_out: Path | None = None
    budget_out: Path | None = None
    # dashboard path resolved up front so the BOM page can link to it
    dash_out: Path | None = None
    if a.dashboard:
        dash_out = (a.dashboard if isinstance(a.dashboard, Path)
                    else a.outdir / f"{stem}_Dashboard.html")

    def _sibling(target: Path | None, of: Path) -> str:
        return (target.name if target is not None
                and target.resolve().parent == of.resolve().parent else "")
    if a.xlsx or a.both:
        xlsx_out = a.xlsx if isinstance(a.xlsx, Path) else a.outdir / f"{stem}_BOM.xlsx"
        write_excel(root, cfg, xlsx_out, warnings, source_rev=a.source_rev,
                    historical=a.historical, src_name=a.input.name,
                    source_url=a.source_url)
        did = True
        if not a.quiet:
            print(f"wrote {xlsx_out}")
    if a.html or a.both:
        p = a.html if isinstance(a.html, Path) else a.outdir / f"{stem}_BOM.html"
        href = a.xlsx_url
        if href is None:
            # sibling .xlsx from the same run -> relative link survives any
            # hosting root (GitHub Pages, GitLab Pages, file://, S3, ...)
            href = _sibling(xlsx_out, p)
        write_html(root, cfg, a.input.name, p, warnings, xlsx_href=href,
                  source_rev=a.source_rev, historical=a.historical,
                  dashboard_href=_sibling(dash_out, p),
                  source_url=a.source_url)
        html_out = p
        did = True
        if not a.quiet:
            print(f"wrote {p}")
    if a.budget:
        budget_out = (a.budget if isinstance(a.budget, Path)
                      else a.outdir / f"{stem}_Budget.xlsx")
        write_budget_xlsx(rollup, cfg, a.input.name, budget_out, warnings,
                          source_rev=a.source_rev, historical=a.historical,
                          source_url=a.source_url)
        did = True
        if not a.quiet:
            print(f"wrote {budget_out}")
    if a.dashboard:
        # relative sibling links, same rule as the BOM page's xlsx button
        write_dashboard(rollup, cfg, a.input.name, dash_out, warnings,
                        budget_href=_sibling(budget_out, dash_out),
                        bom_href=_sibling(html_out, dash_out),
                        source_rev=a.source_rev, historical=a.historical,
                        source_url=a.source_url)
        did = True
        if not a.quiet:
            print(f"wrote {dash_out}")
    if not did:
        print("bomgen: nothing to do (use --xlsx, --html, --both, --budget, "
              "or --dashboard)", file=sys.stderr)
        return 2
    return 0


if __name__ == "__main__":
    sys.exit(main())
