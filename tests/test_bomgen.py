# SPDX-License-Identifier: MIT
# SPDX-FileCopyrightText: 2026 E Douglas, University of Arizona and contributors
"""Tests for bomgen. Run: python -m pytest tests/ -v"""
import json
import sys
from pathlib import Path

import pytest

REPO = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(REPO))
import bomgen  # noqa: E402

FIX = Path(__file__).parent / "fixtures"
SAMPLE_CSV = REPO / "examples" / "NCC-1701_pdmout.csv"


def parse(path, cfg=None):
    cfg = cfg or bomgen.load_config(None)
    warnings = []
    if path.suffix == ".xml":
        header, rows = bomgen.read_xml(path, cfg, warnings)
    else:
        header, rows = bomgen.read_csv(path)
    root = bomgen.build_tree(header, rows, cfg, warnings)
    bomgen.derive(root, cfg)
    bomgen.enrich_materials(root, cfg, warnings)
    return root, warnings


# ------------------------------------------------------------------ CSV path

def test_csv_tree_shape():
    root, _ = parse(SAMPLE_CSV)
    nodes = bomgen.preorder(root)
    assert len(nodes) == 49  # 48 rows + synthesized root
    assert max(n.depth for n in nodes) == 4
    idx = {n.path: n for n in nodes}
    assert idx["0"].part_number  # synthesized root populated from config


def test_excel_float_mangle_repair():
    root, warnings = parse(SAMPLE_CSV)
    idx = {n.path: n for n in bomgen.preorder(root)}
    assert "2.10" in idx, "duplicate '2.1' should be repaired to '2.10'"
    assert any(w.startswith("R1") for w in warnings)


def test_qty_rollup_multiplies_ancestors():
    root, _ = parse(SAMPLE_CSV)
    idx = {n.path: n for n in bomgen.preorder(root)}
    n = idx["3.1.12.2"]
    assert n.qty == 1 and idx["3.1.12"].qty == 6 and n.qty_total == 6


def test_rollup_invariant_leaf_sums():
    root, _ = parse(SAMPLE_CSV)
    total = sum(n.qty_total for n in bomgen.preorder(root) if not n.children)
    root2, _ = parse(SAMPLE_CSV)
    total2 = sum(n.qty_total for n in bomgen.preorder(root2) if not n.children)
    assert total == total2 > 0


def test_part_number_rev_dash():
    root, _ = parse(SAMPLE_CSV)
    idx = {n.path: n for n in bomgen.preorder(root)}
    assert idx["1.1.1"].part_number == "NCC-FP005-02"  # Rev 2 -> -02 (D3/O1)


def test_golden_paths(tmp_path):
    """Snapshot of (path, qty_total) pairs; update deliberately if rules change."""
    golden_file = FIX / "golden_tree.json"
    root, _ = parse(SAMPLE_CSV)
    actual = {n.path: n.qty_total for n in bomgen.preorder(root)}
    if not golden_file.exists():  # first run generates it
        golden_file.write_text(json.dumps(actual, indent=1, sort_keys=True))
        pytest.skip("golden file created; re-run")
    assert actual == json.loads(golden_file.read_text())


# ------------------------------------------------------------------ XML path

def test_xml_nested_references():
    root, warnings = parse(FIX / "sample_export.xml")
    idx = {n.path: n for n in bomgen.preorder(root)}
    # top doc -> "1"; its two references -> 1.1, 1.2; grandchildren -> 1.1.x
    assert idx["1"].filename == "NCC-FA014.SLDASM"
    assert idx["1.1.2"].qty == 2
    assert idx["1.1.2"].qty_total == 2
    assert idx["1.2"].qty == 4
    assert idx["1.1.2"].cots == "X"          # COTS attr flows through (D1)
    assert idx["1.1.1"].part_number == "NCC-FP005-02"
    assert any("O2" in w for w in warnings)  # verification flag surfaces


# ------------------------------------------------------------------ outputs

def test_outputs_write(tmp_path):
    cfg = bomgen.load_config(None)
    root, warnings = parse(SAMPLE_CSV, cfg)
    x, h = tmp_path / "o.xlsx", tmp_path / "o.html"
    bomgen.write_excel(root, cfg, x)
    bomgen.write_html(root, cfg, "src.csv", h, warnings)
    assert x.stat().st_size > 4000
    page = h.read_text(encoding="utf-8")
    assert "__TITLE__" not in page and '"2.10"' in page

    import openpyxl
    ws = openpyxl.load_workbook(x).active
    assert ws["B16"].value == "Assembly Level"
    row17 = [c.value for c in ws[17] if c.value is not None]
    assert "Qty (Total)" in row17


def test_unresolved_swprop_warning():
    """V7: SW-Mass@.SLDPRT-style cells are flagged, grouped per column."""
    _, warnings = parse(SAMPLE_CSV)
    v7 = [w for w in warnings if w.startswith("V7")]
    assert len(v7) == 1 and "'Mass'" in v7[0] and "SW-Mass@" in v7[0]
    assert "4 row(s)" in v7[0]


def test_warning_banner_in_both_outputs(tmp_path):
    """D6: caught gotchas render in a yellow banner atop Excel and HTML."""
    cfg = bomgen.load_config(None)
    root, warnings = parse(SAMPLE_CSV, cfg)
    assert any(w.startswith("R1") for w in warnings)  # float-mangle caught
    x, h = tmp_path / "o.xlsx", tmp_path / "o.html"
    bomgen.write_excel(root, cfg, x, warnings)
    bomgen.write_html(root, cfg, "src.csv", h, warnings)

    page = h.read_text(encoding="utf-8")
    assert 'id="warnbox"' in page
    assert "SW-Mass@" in page and "R1" in page  # both gotchas surfaced

    import openpyxl
    ws = openpyxl.load_workbook(x).active
    banner = ws["B2"].value
    assert banner and "DATA QUALITY" in banner
    assert "SW-Mass@" in banner and "R1" in banner
    assert ws["B6"].value == "Bill of Materials (BOM)"  # template rows intact


def test_no_banner_on_clean_input(tmp_path):
    """No warnings -> no banner in either output."""
    clean = tmp_path / "clean.csv"
    clean.write_text("Level,Qty,Number,COTS\n"
                     "1,1,NCC-FA002.SLDASM,\n"
                     "1.1,2,NCC-FP001.SLDPRT,X\n", encoding="utf-8")
    cfg = bomgen.load_config(None)
    root, warnings = parse(clean, cfg)
    assert warnings == []
    x, h = tmp_path / "o.xlsx", tmp_path / "o.html"
    bomgen.write_excel(root, cfg, x, warnings)
    bomgen.write_html(root, cfg, "clean.csv", h, warnings)
    assert 'id="warnbox"' not in h.read_text(encoding="utf-8")
    import openpyxl
    assert openpyxl.load_workbook(x).active["B2"].value is None


def test_html_xlsx_download_link(tmp_path):
    """D5: explicit xlsx_href lands in the download button, escaped."""
    cfg = bomgen.load_config(None)
    root, warnings = parse(SAMPLE_CSV, cfg)
    h = tmp_path / "o.html"
    bomgen.write_html(root, cfg, "src.csv", h, warnings, xlsx_href="o.xlsx")
    page = h.read_text(encoding="utf-8")
    assert 'id="dl"' in page and 'href="o.xlsx"' in page
    # no href known -> empty attribute (button removes itself client-side)
    bomgen.write_html(root, cfg, "src.csv", h, warnings)
    assert 'href=""' in h.read_text(encoding="utf-8")


def test_cli_both_links_sibling_xlsx(tmp_path):
    """--both into one directory -> button links the .xlsx by filename,
    the layout scripts/build_pages.sh deploys to GitHub/GitLab Pages."""
    rc = bomgen.main([str(SAMPLE_CSV), "--both", "-o", str(tmp_path), "--quiet"])
    assert rc == 0
    page = (tmp_path / "NCC-1701_pdmout_BOM.html").read_text(encoding="utf-8")
    assert 'href="NCC-1701_pdmout_BOM.xlsx"' in page
    assert (tmp_path / "NCC-1701_pdmout_BOM.xlsx").exists()


def test_cli_xlsx_url_override(tmp_path):
    rc = bomgen.main([str(SAMPLE_CSV), "--html", str(tmp_path / "o.html"),
                      "--xlsx-url", "https://example.com/bom.xlsx", "--quiet"])
    assert rc == 0
    page = (tmp_path / "o.html").read_text(encoding="utf-8")
    assert 'href="https://example.com/bom.xlsx"' in page


def test_source_rev_in_both_outputs(tmp_path):
    """D8: --source-rev (e.g. the vault repo's git hash for the CSV) is
    embedded in both outputs for provenance; omitted -> no trace of it."""
    cfg = bomgen.load_config(None)
    root, warnings = parse(SAMPLE_CSV, cfg)
    x, h = tmp_path / "o.xlsx", tmp_path / "o.html"
    bomgen.write_excel(root, cfg, x, warnings, source_rev="a1b2c3d")
    bomgen.write_html(root, cfg, "src.csv", h, warnings, source_rev="a1b2c3d")

    page = h.read_text(encoding="utf-8")
    assert "a1b2c3d" in page and "rev" in page

    import openpyxl
    ws = openpyxl.load_workbook(x).active
    assert "a1b2c3d" in (ws["B8"].value or "")
    assert ws["B6"].value == "Bill of Materials (BOM)"  # template rows intact

    # omitted -> row 8 stays blank, no stray placeholder text in the HTML
    h2 = tmp_path / "o2.html"
    bomgen.write_html(root, cfg, "src.csv", h2, warnings)
    assert "__SOURCE_REV__" not in h2.read_text(encoding="utf-8")
    x2 = tmp_path / "o2.xlsx"
    bomgen.write_excel(root, cfg, x2, warnings)
    assert openpyxl.load_workbook(x2).active["B8"].value is None


def test_cli_source_rev_flag(tmp_path):
    rc = bomgen.main([str(SAMPLE_CSV), "--html", str(tmp_path / "o.html"),
                      "--source-rev", "deadbee", "--quiet"])
    assert rc == 0
    assert "deadbee" in (tmp_path / "o.html").read_text(encoding="utf-8")


# -------------------------------------------------------- source-url / provenance

SOURCE_URL = "https://github.com/owner/repo/blob/abc1234/src.csv"


def test_source_url_hyperlink_in_html(tmp_path):
    """source_url turns the inline source filename into an <a> hyperlink;
    omitted -> plain <code> (no dead link)."""
    cfg = bomgen.load_config(None)
    root, warnings = parse(SAMPLE_CSV, cfg)
    h = tmp_path / "o.html"
    bomgen.write_html(root, cfg, "src.csv", h, warnings,
                      provenance={"source_url": SOURCE_URL})
    page = h.read_text(encoding="utf-8")
    assert f'href="{SOURCE_URL}"' in page
    assert "<code>src.csv</code>" in page   # filename still in <code>

    # no source_url -> no <a>, bare <code>
    h2 = tmp_path / "o2.html"
    bomgen.write_html(root, cfg, "src.csv", h2, warnings)
    page2 = h2.read_text(encoding="utf-8")
    assert SOURCE_URL not in page2
    assert "<code>src.csv</code>" in page2  # plain, no link


def test_source_url_hyperlink_in_dashboard(tmp_path):
    """source_url threads through write_dashboard."""
    f = tmp_path / "spec.csv"
    f.write_text(SPEC_CSV, encoding="utf-8")
    cfgf = tmp_path / "cfg.toml"
    cfgf.write_text('[columns]\nspecs = "Specs"\n', encoding="utf-8")
    rc = bomgen.main([str(f), "-c", str(cfgf), "--dashboard",
                      "-o", str(tmp_path),
                      "--source-url", SOURCE_URL, "--quiet"])
    assert rc == 0
    dash = (tmp_path / "spec_Dashboard.html").read_text(encoding="utf-8")
    assert f'href="{SOURCE_URL}"' in dash


def test_source_url_hyperlink_in_excel(tmp_path):
    """source_url wires a hyperlink onto the source cell in the BOM xlsx."""
    cfg = bomgen.load_config(None)
    root, warnings = parse(SAMPLE_CSV, cfg)
    x = tmp_path / "o.xlsx"
    bomgen.write_excel(root, cfg, x, warnings,
                       provenance={"source": "src.csv",
                                   "source_url": SOURCE_URL})
    import openpyxl
    ws = openpyxl.load_workbook(x).active
    # row 8 carries the source name + hyperlink
    assert ws["B8"].value and "src.csv" in ws["B8"].value
    assert ws["B8"].hyperlink and SOURCE_URL in ws["B8"].hyperlink.target

    # omitted source_url -> no hyperlink on row 8
    x2 = tmp_path / "o2.xlsx"
    bomgen.write_excel(root, cfg, x2, warnings,
                       provenance={"source": "src.csv"})
    ws2 = openpyxl.load_workbook(x2).active
    assert ws2["B8"].hyperlink is None


def test_source_url_no_hyperlink_when_empty(tmp_path):
    """Empty source_url -> no hyperlink attr, no dead <a> in HTML."""
    cfg = bomgen.load_config(None)
    root, warnings = parse(SAMPLE_CSV, cfg)
    h = tmp_path / "o.html"
    bomgen.write_html(root, cfg, "src.csv", h, warnings,
                      provenance={"source_url": ""})
    page = h.read_text(encoding="utf-8")
    # source rendered as plain <code> without any wrapping <a> link
    assert "<code>src.csv</code>" in page
    # a linked source would look like href="..."><code>src.csv</code>; must be absent
    assert '"><code>src.csv</code>' not in page


def test_source_url_in_budget_xlsx(tmp_path):
    """source_url sets a hyperlink on the Budget sheet meta cell."""
    rollup, warnings = _spec_rollup(tmp_path)
    cfg = bomgen.load_config(None)
    out = tmp_path / "budget.xlsx"
    bomgen.write_budget_xlsx(rollup, cfg, "spec.csv", out, warnings,
                             provenance={"source_url": SOURCE_URL})
    import openpyxl
    wb = openpyxl.load_workbook(out)
    bs = wb["Budget"]
    # row 2 is the meta cell; it should carry the hyperlink
    assert bs["A2"].hyperlink and SOURCE_URL in bs["A2"].hyperlink.target

    # omitted -> no hyperlink
    out2 = tmp_path / "budget2.xlsx"
    bomgen.write_budget_xlsx(rollup, cfg, "spec.csv", out2, warnings)
    bs2 = openpyxl.load_workbook(out2)["Budget"]
    assert bs2["A2"].hyperlink is None


def test_cli_source_url_flag(tmp_path):
    """--source-url propagates to HTML and xlsx via the CLI."""
    rc = bomgen.main([str(SAMPLE_CSV), "--both",
                      "--source-url", SOURCE_URL,
                      "-o", str(tmp_path), "--quiet"])
    assert rc == 0
    page = (tmp_path / "NCC-1701_pdmout_BOM.html").read_text(encoding="utf-8")
    assert f'href="{SOURCE_URL}"' in page
    import openpyxl
    ws = openpyxl.load_workbook(
        tmp_path / "NCC-1701_pdmout_BOM.xlsx").active
    assert ws["B8"].hyperlink and SOURCE_URL in ws["B8"].hyperlink.target


def test_provenance_details_block_in_html(tmp_path):
    """A collapsible <details> provenance block is embedded in the HTML header."""
    cfg = bomgen.load_config(None)
    root, warnings = parse(SAMPLE_CSV, cfg)
    h = tmp_path / "o.html"
    bomgen.write_html(root, cfg, "src.csv", h, warnings, source_rev="abc1234",
                      provenance={"source": "src.csv",
                                  "source_url": SOURCE_URL,
                                  "source_rev": "abc1234"})
    page = h.read_text(encoding="utf-8")
    # block present with key identifiers
    assert '<details id="prov">' in page
    assert "Build provenance" in page
    assert SOURCE_URL in page          # source URL in the details body
    assert "abc1234" in page           # source rev
    assert bomgen.__version__ in page


def test_provenance_details_block_in_dashboard(tmp_path):
    """Dashboard HTML also carries the collapsible provenance block."""
    f = tmp_path / "spec.csv"
    f.write_text(SPEC_CSV, encoding="utf-8")
    cfgf = tmp_path / "cfg.toml"
    cfgf.write_text('[columns]\nspecs = "Specs"\n', encoding="utf-8")
    rc = bomgen.main([str(f), "-c", str(cfgf), "--dashboard",
                      "-o", str(tmp_path),
                      "--source-url", SOURCE_URL,
                      "--source-rev", "def5678", "--quiet"])
    assert rc == 0
    dash = (tmp_path / "spec_Dashboard.html").read_text(encoding="utf-8")
    assert '<details id="prov">' in dash
    assert SOURCE_URL in dash
    assert "def5678" in dash


def test_pdm_file_url_links(tmp_path):
    """Configurable file_url_template turns filenames into PDM-viewer links
    in both outputs; {file} is substituted with the URL-encoded filename."""
    cfg = bomgen.load_config(None)
    cfg["links"]["file_url_template"] = (
        "https://pdm.example.edu/vault/PROJ?view=bom&file={file}")
    root, warnings = parse(SAMPLE_CSV, cfg)
    idx = {n.path: n for n in bomgen.preorder(root)}
    # a known part row: NCC-FP005.SLDPRT
    node = idx["1.1.1"]
    assert node.file_url == (
        "https://pdm.example.edu/vault/PROJ?view=bom&file=NCC-FP005.SLDPRT")
    # a filename with spaces is URL-encoded
    shim = idx["1.1.5"]  # "Custom Injector Shim_WRP.SLDPRT"
    assert "%20" in shim.file_url and " " not in shim.file_url

    x, h = tmp_path / "o.xlsx", tmp_path / "o.html"
    bomgen.write_excel(root, cfg, x, warnings)
    bomgen.write_html(root, cfg, "src.csv", h, warnings)

    page = h.read_text(encoding="utf-8")
    # the per-row fileUrl lands in the embedded JSON data blob
    assert '"fileUrl": "https://pdm.example.edu' in page
    assert "file=NCC-FP005.SLDPRT" in page

    import openpyxl
    ws = openpyxl.load_workbook(x).active
    # find the Part Name cell for the NCC-FP005 row and check its hyperlink
    links = [c.hyperlink.target for row in ws.iter_rows() for c in row
             if c.hyperlink]
    assert any("NCC-FP005.SLDPRT" in t for t in links)


def test_pdm_file_url_default_off(tmp_path):
    """Default config (empty template) -> no file_url, no links; unchanged."""
    cfg = bomgen.load_config(None)
    assert cfg["links"]["file_url_template"] == ""
    root, warnings = parse(SAMPLE_CSV, cfg)
    assert all(n.file_url == "" for n in bomgen.preorder(root))
    h = tmp_path / "o.html"
    bomgen.write_html(root, cfg, "src.csv", h, warnings)
    page = h.read_text(encoding="utf-8")
    # fileUrl field is emitted but empty for every row; no populated links
    assert '"fileUrl": ""' in page and '"fileUrl": "http' not in page


def test_pdm_found_in_url_auto():
    """By default {found_in} strips the leading <drive>:\\<vault>\\ for ANY
    drive letter / vault name — no found_in_strip needed."""
    cfg = bomgen.load_config(None)
    cfg["links"]["file_url_template"] = (
        "https://pdm.example.edu/solidworkspdm/Starfleet_PDM/{found_in}")
    root, _ = parse(SAMPLE_CSV, cfg)
    idx = {n.path: n for n in bomgen.preorder(root)}
    # 1.1.1 Found In: D:\Starfleet_PDM\SFC\ENT\NCC\Flight -> drive+vault dropped
    assert idx["1.1.1"].file_url == (
        "https://pdm.example.edu/solidworkspdm/Starfleet_PDM/SFC/ENT/NCC/Flight")
    # a folder with spaces (…\WIP\9000 - Piece Parts) is percent-encoded
    assert "9000%20-%20Piece%20Parts" in idx["1.1.3"].file_url
    assert " " not in idx["1.1.3"].file_url


def test_pdm_found_in_drive_agnostic():
    """Any drive letter works; the vault root is dropped by name-position."""
    rest = bomgen._found_in_rest  # helper: (found_in, strip) -> url tail
    assert rest("D:\\Obs_PDM\\STP\\ESC\\Flight", "") == "STP/ESC/Flight"
    assert rest("E:\\Obs_PDM\\STP\\ESC\\Flight", "") == "STP/ESC/Flight"
    assert rest("Z:\\Another_Vault\\A\\B", "") == "A/B"
    # explicit override still supported (drive-tolerant)
    assert rest("C:\\V\\STP\\ESC", "C:\\V\\STP") == "ESC"
    assert rest("C:\\V\\STP\\ESC", "V\\STP") == "ESC"


def test_pdm_found_in_url():
    """Explicit found_in_strip override yields the same tail."""
    cfg = bomgen.load_config(None)
    cfg["links"]["file_url_template"] = (
        "https://pdm.example.edu/solidworkspdm/Starfleet_PDM/{found_in}")
    cfg["links"]["found_in_strip"] = "D:\\Starfleet_PDM"
    root, _ = parse(SAMPLE_CSV, cfg)
    idx = {n.path: n for n in bomgen.preorder(root)}
    assert idx["1.1.1"].file_url == (
        "https://pdm.example.edu/solidworkspdm/Starfleet_PDM/SFC/ENT/NCC/Flight")


def test_pdm_found_in_and_file():
    """Both placeholders combine: folder path + filename query."""
    cfg = bomgen.load_config(None)
    cfg["links"]["file_url_template"] = (
        "https://pdm.example.edu/vault/{found_in}?view=bom&file={file}")
    cfg["links"]["found_in_strip"] = "D:\\Starfleet_PDM"
    root, _ = parse(SAMPLE_CSV, cfg)
    node = {n.path: n for n in bomgen.preorder(root)}["1.1.1"]
    assert node.file_url == (
        "https://pdm.example.edu/vault/SFC/ENT/NCC/Flight"
        "?view=bom&file=NCC-FP005.SLDPRT")


# ------------------------------------------------------------ budget / specs

SPEC_CSV = ("Level,Qty,Number,Name,Rev,Specs,COTS\n"
            "1,1,NCC-FA010.SLDASM,Weldment,1,SPEC-WELD-001,\n"
            "1.1,4,NCC-FP020.SLDPRT,Plate,,,\n"
            "2,3,NCC-FP030.SLDPRT,Bracket,2,SPEC-MACH-002,\n"
            "3,1,NCC-FA011.SLDASM,Mount,,,\n"
            "3.1,6,92423A111_Screw.SLDPRT,,,SPEC-HDWE-003,X\n"
            "3.2,1,NCC-FP032.SLDPRT,Shim,,,\n"
            "4,2,NCC-FP030.SLDPRT,Bracket,2,SPEC-MACH-002,\n"
            "5,1,NCC-FP040.SLDPRT,MultiSpec,,SPEC-MACH-002; SPEC-WELD-001,\n")


def _spec_rollup(tmp_path, csv_text=SPEC_CSV, cfg=None):
    f = tmp_path / "spec.csv"
    f.write_text(csv_text, encoding="utf-8")
    cfg = cfg or bomgen.load_config(None)
    # the repo's own bomgen.toml remaps specs for the demo site; these
    # fixtures use a literal "Specs" column, so pin the mapping
    cfg["columns"]["specs"] = "Specs"
    warnings = []
    header, rows = bomgen.read_csv(f)
    root = bomgen.build_tree(header, rows, cfg, warnings)
    bomgen.derive(root, cfg)
    rollup = bomgen.budget_rollup(root, cfg, warnings, header)
    return rollup, warnings


def test_budget_rollup_semantics(tmp_path):
    """Spec'd node covers its subtree; same part under one spec merges with
    summed qty; leafs with no coverage land in unassigned; multi-spec uses
    the first listed spec."""
    rollup, warnings = _spec_rollup(tmp_path)
    by_name = {s["name"]: s for s in rollup["specs"]}
    assert set(by_name) == {"SPEC-WELD-001", "SPEC-MACH-002", "SPEC-HDWE-003"}
    # weldment is one line; its Plate child is covered, NOT unassigned
    weld = by_name["SPEC-WELD-001"]["lines"]
    assert len(weld) == 1 and weld[0]["file"] == "NCC-FA010.SLDASM"
    # bracket appears twice in the tree (qty 3 + 2) -> one merged line
    mach = {l["file"]: l for l in by_name["SPEC-MACH-002"]["lines"]}
    assert mach["NCC-FP030.SLDPRT"]["qty"] == 5
    assert mach["NCC-FP030.SLDPRT"]["occ"] == 2
    # multi-spec part landed under its FIRST spec
    assert "NCC-FP040.SLDPRT" in mach
    # shim is the only unassigned part
    ufiles = [l["file"] for l in rollup["unassigned"]["lines"]]
    assert ufiles == ["NCC-FP032.SLDPRT"]
    assert rollup["counts"]["specs"] == 3
    assert any(w.startswith("V10") for w in warnings)   # unassigned shim
    assert any(w.startswith("V11") for w in warnings)   # multi-spec row


def test_budget_specs_column_missing_warns(tmp_path):
    csv = "Level,Qty,Number\n1,1,NCC-FA001.SLDASM\n1.1,2,NCC-FP001.SLDPRT\n"
    rollup, warnings = _spec_rollup(tmp_path, csv_text=csv)
    assert rollup["counts"]["specs"] == 0
    assert any(w.startswith("V10") and "not in input" in w for w in warnings)


def test_budget_nested_spec_warns(tmp_path):
    csv = ("Level,Qty,Number,Specs\n"
           "1,1,NCC-FA010.SLDASM,SPEC-WELD-001\n"
           "1.1,2,NCC-FP099.SLDPRT,SPEC-HDWE-003\n")
    rollup, warnings = _spec_rollup(tmp_path, csv_text=csv)
    assert any(w.startswith("V12") for w in warnings)
    # nested item still gets its own line under its own spec
    names = {s["name"] for s in rollup["specs"]}
    assert names == {"SPEC-WELD-001", "SPEC-HDWE-003"}


def test_budget_spec_url_template(tmp_path):
    cfg = bomgen.load_config(None)
    cfg["links"]["spec_url_template"] = "https://docs.example.edu/{spec}"
    rollup, _ = _spec_rollup(tmp_path, cfg=cfg)
    urls = {s["name"]: s["url"] for s in rollup["specs"]}
    assert urls["SPEC-MACH-002"] == "https://docs.example.edu/SPEC-MACH-002"


def test_budget_workbook_structure(tmp_path):
    rollup, warnings = _spec_rollup(tmp_path)
    cfg = bomgen.load_config(None)
    out = tmp_path / "budget.xlsx"
    bomgen.write_budget_xlsx(rollup, cfg, "spec.csv", out, warnings,
                             source_rev="abc1234")
    import openpyxl
    wb = openpyxl.load_workbook(out)
    assert wb.sheetnames == ["Budget", "Parts"]
    ps = wb["Parts"]
    rows = [r for r in ps.iter_rows(min_row=2, max_col=6, values_only=True)
            if r[0]]
    # merged bracket line: qty 5 under SPEC-MACH-002
    bracket = [r for r in rows if r[1] == "NCC-FP030-02"]
    assert bracket and bracket[0][0] == "SPEC-MACH-002" and bracket[0][5] == 5
    # unit-best + ext formulas present, and Budget sheet SUMIF-links Parts
    assert ps["J2"].value.startswith("=IF(I2")
    assert ps["N2"].value.startswith("=IF(J2")
    bs = wb["Budget"]
    cells = [c.value for row in bs.iter_rows() for c in row
             if isinstance(c.value, str)]
    assert any("SUMIF(Parts!$A:$A" in v for v in cells)
    assert any(v == "TOTAL" for v in cells)
    assert any("abc1234" in v for v in cells)           # provenance stamp


def test_cli_budget_dashboard_outputs(tmp_path):
    f = tmp_path / "spec.csv"
    f.write_text(SPEC_CSV, encoding="utf-8")
    # explicit config: the repo's own bomgen.toml (picked up from cwd by
    # default) remaps specs for the demo site
    cfgf = tmp_path / "cfg.toml"
    cfgf.write_text('[columns]\nspecs = "Specs"\n', encoding="utf-8")
    rc = bomgen.main([str(f), "-c", str(cfgf), "--both", "--budget",
                      "--dashboard", "-o", str(tmp_path), "--quiet"])
    assert rc == 0
    dash = (tmp_path / "spec_Dashboard.html").read_text(encoding="utf-8")
    assert (tmp_path / "spec_Budget.xlsx").exists()
    # dashboard links its xlsx twin and the BOM page (relative siblings)
    assert 'href="spec_Budget.xlsx"' in dash
    assert 'href="spec_BOM.html"' in dash
    # rollup data embedded; warnings banner present
    assert "SPEC-MACH-002" in dash and 'id="warnbox"' in dash
    # and the BOM page links forward to the dashboard (discoverability)
    bom = (tmp_path / "spec_BOM.html").read_text(encoding="utf-8")
    assert 'id="dash"' in bom and 'href="spec_Dashboard.html"' in bom


def test_pdm_found_in_missing_skips_link(tmp_path):
    """A {found_in} template + a row with no Found In -> no dead link."""
    csv = tmp_path / "nofoundin.csv"
    csv.write_text("Level,Qty,Number,Found In\n"
                   "1,1,NCC-FA002.SLDASM,\n"
                   "1.1,1,NCC-FP001.SLDPRT,\n", encoding="utf-8")
    cfg = bomgen.load_config(None)
    cfg["links"]["file_url_template"] = "https://pdm.example.edu/v/{found_in}"
    root, _ = parse(csv, cfg)
    assert all(n.file_url == "" for n in bomgen.preorder(root))


# ------------------------------------------------------------ diff / historical

DIFF_PREV = ("Level,Qty,Number,Name,Specs,Checked Out By\n"
             "1,1,NCC-FA010.SLDASM,Weldment,SPEC-A,\n"
             "1.1,2,NCC-FP030.SLDPRT,Bracket,,\n"
             "2,1,NCC-FP050.SLDPRT,Gasket,,alice\n")
DIFF_CUR = ("Level,Qty,Number,Name,Specs,Checked Out By\n"
            "1,1,NCC-FA010.SLDASM,Weldment,SPEC-A,\n"
            "1.1,3,NCC-FP030.SLDPRT,Bracket,,\n"
            "3,1,NCC-FP060.SLDPRT,NewPart,,bob\n")


def _diffed(tmp_path, prev_text, cur_text, cfg=None):
    prev = tmp_path / "prev.csv"; prev.write_text(prev_text, encoding="utf-8")
    cur = tmp_path / "cur.csv"; cur.write_text(cur_text, encoding="utf-8")
    cfg = cfg or bomgen.load_config(None)
    warnings = []
    header, rows = bomgen.read_csv(cur)
    root = bomgen.build_tree(header, rows, cfg, warnings)
    bomgen.derive(root, cfg)
    bomgen.apply_diff(root, prev, cfg, warnings)
    return root, warnings


def test_diff_changed_added_removed(tmp_path):
    root, warnings = _diffed(tmp_path, DIFF_PREV, DIFF_CUR)
    idx = {n.filename: n for n in bomgen.preorder(root) if n.filename}
    assert idx["NCC-FP030.SLDPRT"].diff == "changed"   # qty 2 -> 3
    assert idx["NCC-FP060.SLDPRT"].diff == "added"
    assert idx["NCC-FA010.SLDASM"].diff == ""          # untouched
    d = [w for w in warnings if w.startswith("DIFF")]
    assert len(d) == 1 and "1 row(s) changed" in d[0] and "1 added" in d[0]
    assert "NCC-FP050.SLDPRT" in d[0]                  # removed, in banner


def test_diff_ignores_unmapped_and_configured_columns(tmp_path):
    # only "Checked Out By" (unmapped churn) differs -> no change flagged
    prev = DIFF_CUR.replace(",bob\n", ",carol\n")
    root, warnings = _diffed(tmp_path, prev, DIFF_CUR)
    assert all(n.diff == "" for n in bomgen.preorder(root))
    assert not any(w.startswith("DIFF") for w in warnings)
    # a mapped column can be excluded via [diff].ignore_columns
    cfg = bomgen.load_config(None)
    cfg["diff"]["ignore_columns"] = ["Qty"]
    root, _ = _diffed(tmp_path, DIFF_PREV.replace(",alice\n", ",\n")
                      .replace("2,1,NCC-FP050.SLDPRT,Gasket,,\n", ""),
                      DIFF_CUR.replace("3,1,NCC-FP060.SLDPRT,NewPart,,bob\n", ""),
                      cfg=cfg)
    idx = {n.filename: n for n in bomgen.preorder(root) if n.filename}
    assert idx["NCC-FP030.SLDPRT"].diff == ""          # qty diff ignored


def test_diff_move_is_not_a_change(tmp_path):
    prev = ("Level,Qty,Number,Name\n"
            "1,1,NCC-FA010.SLDASM,Asm\n"
            "1.1,2,NCC-FP030.SLDPRT,Bracket\n")
    cur = ("Level,Qty,Number,Name\n"
           "1,1,NCC-FA010.SLDASM,Asm\n"
           "2,2,NCC-FP030.SLDPRT,Bracket\n")   # moved to top level, same data
    root, warnings = _diffed(tmp_path, prev, cur)
    idx = {n.filename: n for n in bomgen.preorder(root) if n.filename}
    assert idx["NCC-FP030.SLDPRT"].diff == ""
    assert not any(w.startswith("DIFF") for w in warnings)


def test_diff_unreadable_prev_warns_not_aborts(tmp_path):
    cur = tmp_path / "cur.csv"; cur.write_text(DIFF_CUR, encoding="utf-8")
    cfg = bomgen.load_config(None)
    warnings = []
    header, rows = bomgen.read_csv(cur)
    root = bomgen.build_tree(header, rows, cfg, warnings)
    bomgen.derive(root, cfg)
    bomgen.apply_diff(root, tmp_path / "nope.csv", cfg, warnings)
    assert any(w.startswith("DIFF") and "skipped" in w for w in warnings)
    assert all(n.diff == "" for n in bomgen.preorder(root))


def test_cli_diff_and_historical(tmp_path):
    prev = tmp_path / "prev.csv"; prev.write_text(DIFF_PREV, encoding="utf-8")
    cur = tmp_path / "cur.csv"; cur.write_text(DIFF_CUR, encoding="utf-8")
    rc = bomgen.main([str(cur), "--both", "--budget", "--dashboard",
                      "--diff-against", str(prev),
                      "--historical", "v0.2 (2026-07-01)",
                      "-o", str(tmp_path), "--quiet"])
    assert rc == 0
    page = (tmp_path / "cur_BOM.html").read_text(encoding="utf-8")
    assert 'id="histbar"' in page and "Historical version v0.2" in page
    assert '"diff": "changed"' in page and '"diff": "added"' in page
    dash = (tmp_path / "cur_Dashboard.html").read_text(encoding="utf-8")
    assert 'id="histbar"' in dash and '"diffed": true' in dash

    import openpyxl
    ws = openpyxl.load_workbook(tmp_path / "cur_BOM.xlsx").active
    assert "HISTORICAL VERSION" in (ws["B1"].value or "")
    # some data cell carries the green diff fill
    greens = [c for row in ws.iter_rows() for c in row
              if c.fill and c.fill.fgColor and c.fill.fgColor.rgb == "00E2EFDA"]
    assert greens
    wb = openpyxl.load_workbook(tmp_path / "cur_Budget.xlsx")
    bcells = [c.value for row in wb["Budget"].iter_rows() for c in row
              if isinstance(c.value, str)]
    assert any("HISTORICAL VERSION" in v for v in bcells)
    pgreens = [c for row in wb["Parts"].iter_rows() for c in row
               if c.fill and c.fill.fgColor and c.fill.fgColor.rgb == "00E2EFDA"]
    assert pgreens


def test_build_pages_history_integration(tmp_path):
    """End-to-end: a git repo with two tags -> per-tag pages, versions.js,
    yellow chrome on tags only, skip-safe."""
    import subprocess, os
    repo = tmp_path / "repo"; repo.mkdir()
    env = {**os.environ, "PYTHONPATH": str(REPO), "BUILD_HISTORY": "1",
           "BOMGEN": "python -m bomgen",
           "GIT_AUTHOR_NAME": "t", "GIT_AUTHOR_EMAIL": "t@e.c",
           "GIT_COMMITTER_NAME": "t", "GIT_COMMITTER_EMAIL": "t@e.c"}
    def git(*args):
        subprocess.run(["git", *args], cwd=repo, env=env, check=True,
                       capture_output=True)
    git("init", "-q")
    (repo / "bom.csv").write_text(DIFF_PREV, encoding="utf-8")
    git("add", "-A"); git("commit", "-qm", "one"); git("tag", "t1")
    (repo / "bom.csv").write_text(DIFF_CUR, encoding="utf-8")
    git("add", "-A"); git("commit", "-qm", "two"); git("tag", "t2")
    r = subprocess.run(["bash", str(REPO / "scripts" / "build_pages.sh"),
                        "bom.csv", str(REPO / "bomgen.toml"), "site"],
                       cwd=repo, env=env, capture_output=True, text=True)
    assert r.returncode == 0, r.stderr
    site = repo / "site"
    assert (site / "index.html").exists()
    assert (site / "v" / "t1" / "index.html").exists()
    assert (site / "v" / "t2" / "index.html").exists()
    assert "t2" in (site / "versions.js").read_text()
    t2 = (site / "v" / "t2" / "index.html").read_text(encoding="utf-8")
    assert 'id="histbar"' in t2                       # yellow on tag page
    assert '"diff": "changed"' in t2                  # diffed vs t1
    cur = (site / "index.html").read_text(encoding="utf-8")
    assert 'id="histbar"' not in cur                  # current not yellow
    v2 = (site / "v" / "t2" / "versions.js").read_text()
    assert '"../../index.html"' in v2                 # relative root link


# ------------------------------------------------------------ provenance

def test_provenance_in_all_outputs(tmp_path):
    """Build-provenance: repo/branch/commit + linked source path + toolchain
    versions appear as a <details> block on both pages and a cell in both
    workbooks."""
    f = tmp_path / "spec.csv"
    f.write_text(SPEC_CSV, encoding="utf-8")
    cfgf = tmp_path / "cfg.toml"
    cfgf.write_text('[columns]\nspecs = "Specs"\n', encoding="utf-8")
    rc = bomgen.main([str(f), "-c", str(cfgf), "--both", "--budget",
                      "--dashboard", "-o", str(tmp_path), "--quiet",
                      "--repo", "douglase/pdmbomgen", "--branch", "main",
                      "--commit", "abc1234", "--source-rev", "def5678",
                      "--source-path", "vault/spec.csv",
                      "--source-url",
                      "https://github.com/douglase/pdmbomgen/blob/abc1234/vault/spec.csv"])
    assert rc == 0
    import openpyxl, platform
    opx, py = openpyxl.__version__, platform.python_version()

    for page_name in ("spec_BOM.html", "spec_Dashboard.html"):
        page = (tmp_path / page_name).read_text(encoding="utf-8")
        assert '<details id="prov">' in page and "Build provenance" in page
        for frag in ("douglase/pdmbomgen", "main", "abc1234", "def5678",
                     "vault/spec.csv", opx, py, bomgen.__version__):
            assert frag in page, (page_name, frag)
        assert 'href="https://github.com/douglase/pdmbomgen/blob/' in page

    ws = openpyxl.load_workbook(tmp_path / "spec_BOM.xlsx").active
    cell = ws["B14"]
    for frag in ("douglase/pdmbomgen", "abc1234", "vault/spec.csv", opx, py):
        assert frag in cell.value
    assert cell.hyperlink and "blob/abc1234" in cell.hyperlink.target

    wb = openpyxl.load_workbook(tmp_path / "spec_Budget.xlsx")
    cells = [c for row in wb["Budget"].iter_rows() for c in row
             if isinstance(c.value, str) and "openpyxl" in c.value]
    assert cells and "douglase/pdmbomgen" in cells[0].value


def test_provenance_defaults_without_git_flags(tmp_path):
    """No git flags -> block still present with toolchain versions; no repo/
    branch/commit rows and no dead link."""
    f = tmp_path / "plain.csv"
    f.write_text("Level,Qty,Number\n1,1,NCC-FA001.SLDASM\n", encoding="utf-8")
    rc = bomgen.main([str(f), "--html", str(tmp_path / "o.html"), "--quiet"])
    assert rc == 0
    page = (tmp_path / "o.html").read_text(encoding="utf-8")
    assert '<details id="prov">' in page
    import platform
    assert platform.python_version() in page
    assert "Repository" not in page and "Build commit" not in page


# ------------------------------------------------------------ materials (Stage B)

MATERIALS_JSON = FIX / "materials_raw.json"


def _materials_cfg(**over):
    cfg = bomgen.load_config(None)
    cfg["materials"].update({
        "enabled": True, "cache_file": str(MATERIALS_JSON),
        "properties": ["Density_kg/m3", "Tensile_Strength_mpa"],
    })
    cfg["materials"].update(over)
    return cfg


def test_materials_key_normalization():
    k = bomgen.material_cache_key
    assert k("  Tritanium   18-8 ") == k("tritanium 18-8")
    assert k("") == "" and k(None) == ""


def test_materials_enrichment_match():
    """A raw /export/raw-json dump enriches rows whose Material matches."""
    root, warnings = parse(SAMPLE_CSV, _materials_cfg())
    idx = {n.path: n for n in bomgen.preorder(root)}
    # 1.1.3 is a "Tritanium 18-8" washer in the example CSV
    assert idx["1.1.3"].material_props["Density_kg/m3"] == "8000 kg/m3"
    assert idx["1.1.3"].material_props["Tensile_Strength_mpa"] == "620 MPa"
    # 1.1.4 is "Duranium A286 Alloy" — has density, no tensile in the DB
    assert idx["1.1.4"].material_props["Density_kg/m3"] == "7920 kg/m3"
    assert idx["1.1.4"].material_props["Tensile_Strength_mpa"] == ""
    # unmatched materials in the example produce a grouped V9 warning
    assert any(w.startswith("V9") for w in warnings)


def test_materials_show_units_false():
    root, _ = parse(SAMPLE_CSV, _materials_cfg(show_units=False))
    idx = {n.path: n for n in bomgen.preorder(root)}
    assert idx["1.1.3"].material_props["Density_kg/m3"] == "8000"


def test_materials_synonym_match(tmp_path):
    """A Material that matches a DB synonym still resolves."""
    csv = tmp_path / "syn.csv"
    csv.write_text("Level,Qty,Number,Material,COTS\n"
                   "1,1,NCC-FA002.SLDASM,,\n"
                   "1.1,1,NCC-FP001.SLDPRT,18-8,\n", encoding="utf-8")
    root, _ = parse(csv, _materials_cfg())
    idx = {n.path: n for n in bomgen.preorder(root)}
    assert idx["1.1"].material_props["Density_kg/m3"] == "8000 kg/m3"


def test_materials_cache_missing_warns():
    cfg = _materials_cfg(cache_file="/no/such/materials.json")
    root, warnings = parse(SAMPLE_CSV, cfg)
    assert any(w.startswith("V8") for w in warnings)
    assert all(n.material_props.get("Density_kg/m3", "") == ""
               for n in bomgen.preorder(root) if n.parent)


def test_materials_disabled_is_noop():
    """Default config -> no enrichment, no material columns, no warnings."""
    cfg = bomgen.load_config(None)
    assert cfg["materials"]["enabled"] is False
    root, warnings = parse(SAMPLE_CSV, cfg)
    assert all(n.material_props == {} for n in bomgen.preorder(root))
    assert not any(w.startswith(("V8", "V9")) for w in warnings)


def test_materials_render_both_outputs(tmp_path):
    cfg = _materials_cfg(labels={"Density_kg/m3": "Density"})
    root, warnings = parse(SAMPLE_CSV, cfg)
    x, h = tmp_path / "o.xlsx", tmp_path / "o.html"
    bomgen.write_excel(root, cfg, x, warnings)
    bomgen.write_html(root, cfg, "src.csv", h, warnings)

    page = h.read_text(encoding="utf-8")
    assert '"Density"' in page and "8000 kg/m3" in page

    import openpyxl
    ws = openpyxl.load_workbook(x).active
    row17 = [c.value for c in ws[17] if c.value is not None]
    assert "Density" in row17            # labeled header
    assert "Tensile Strength mpa" in row17  # default underscore->space label
